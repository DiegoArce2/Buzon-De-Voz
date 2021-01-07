from flask import Flask, request
from werkzeug.utils import secure_filename
import os
import glob
import os
import librosa
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.pyplot import specgram
import soundfile as sf
import sounddevice as sd
import queue
import time
import time
import argparse
import numpy as np
import keras
from keras.models import Sequential
from keras.layers import Dense, Dropout, Activation
from keras.layers import Dense, Dropout
from keras.layers import Embedding
from keras.layers import Conv1D, GlobalAveragePooling1D, MaxPooling1D
from keras.optimizers import SGD
import os
import os.path as op
from sklearn.model_selection import train_test_split
from random import randint, uniform,random
from os import remove
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook 
from datetime import date,datetime
import pandas.io.formats.excel
import zipfile
import json
from keras import backend as K
import tensorflow as tf
from sklearn.metrics import roc_auc_score
nombreArchivo=""

def verificarModelo(modelo):
    contenido = os.listdir('./data/')    
    for x in contenido:
        print(x)
        if(x.lower()==modelo.lower()):
            return True
    return False


def recall_m(y_true, y_pred):
    true_positives = K.sum(K.round(K.clip(y_true * y_pred, 0, 1)))
    possible_positives = K.sum(K.round(K.clip(y_true, 0, 1)))
    recall = true_positives / (possible_positives + K.epsilon())
    return recall

def precision_m(y_true, y_pred):
    true_positives = K.sum(K.round(K.clip(y_true * y_pred, 0, 1)))
    predicted_positives = K.sum(K.round(K.clip(y_pred, 0, 1)))
    precision = true_positives / (predicted_positives + K.epsilon())
    return precision

def f1_m(y_true, y_pred):
    precision = precision_m(y_true, y_pred)
    recall = recall_m(y_true, y_pred)
    return 2*((precision*recall)/(precision+recall+K.epsilon()))









modelsCargados = []
modelsCargados.append([1,2])
def CargaModelos():
    modelsCargados.clear()
    contenido = os.listdir('./models/')    
    for x in contenido:
        print(x)
        model = keras.models.load_model("./models/"+x+"",custom_objects= {'f1_m': f1_m,'precision_m':precision_m,'recall_m':recall_m})
        y=x.replace(".h5", "")
        modelsCargados.append([y,model])
        
CargaModelos()


def BuscarModelo(mod):
    print("ss"+mod)
    for x in modelsCargados:
        if x[0]=="Model"+str(mod).lower():
            print(x[0])
            return x[1]

        
    





def parse_predict_files(archivo,parent_dir,file_ext='*.wav'):
    features = np.empty((0,168))
    filenames = []
    for fn in glob.glob(os.path.join(parent_dir, file_ext)):
        nombre=fn.replace("predict\\", "")
        if nombre==archivo:
            mfccs,mel= extract_feature(fn)
            ext_features = np.hstack([mfccs,mel])
            features = np.vstack([features,ext_features])
            filenames.append(fn)
            return np.array(features), np.array(filenames)


def detect_leading_silence(sound, silence_threshold=-50.0, chunk_size=10):
    trim_ms = 0 # ms

    assert chunk_size > 0 # to avoid infinite loop
    while sound[trim_ms:trim_ms+chunk_size].dBFS < silence_threshold and trim_ms < len(sound):
        trim_ms += chunk_size

    return trim_ms





start = time.time()

def extract_feature(file_name=None):
    if file_name: 
        print('Extracting', file_name,' Tiempo actual: ',(time.time()-start))
        X, sample_rate = sf.read(file_name, dtype='float32')
    else:  
        device_info = sd.query_devices(None, 'input')
        sample_rate = int(device_info['default_samplerate'])
        q = queue.Queue()
        def callback(i,f,t,s): q.put(i.copy())
        data = []
        with sd.InputStream(samplerate=sample_rate, callback=callback):
            while True: 
                if len(data) < 100000: data.extend(q.get())
                else: break
        X = np.array(data)

    if X.ndim > 1: X = X[:,0]
    X = X.T

    # mfcc (mel-frequency cepstrum)
    mfccs = np.mean(librosa.feature.mfcc(y=X, sr=sample_rate, n_mfcc=40).T,axis=0)

    # melspectrogram
    mel = np.mean(librosa.feature.melspectrogram(X, sr=sample_rate).T,axis=0)

    return mfccs,mel

from pydub import AudioSegment

def detect_leading_silence(sound, silence_threshold=-50.0, chunk_size=10):
    trim_ms = 0 # ms

    assert chunk_size > 0 # to avoid infinite loop
    while sound[trim_ms:trim_ms+chunk_size].dBFS < silence_threshold and trim_ms < len(sound):
        trim_ms += chunk_size

    return trim_ms

from pydub import AudioSegment

def detect_leading_silence(sound, silence_threshold=-50.0, chunk_size=10):
    trim_ms = 0 # ms

    assert chunk_size > 0 # to avoid infinite loop
    while sound[trim_ms:trim_ms+chunk_size].dBFS < silence_threshold and trim_ms < len(sound):
        trim_ms += chunk_size

    return trim_ms

def parse_audio_files(parent_dir,file_ext='*.wav'):
    sub_dirs = os.listdir(parent_dir)
    sub_dirs.sort()
    features, labels = np.empty((0,168)), np.empty(0)
    for label, sub_dir in enumerate(sub_dirs):
        if os.path.isdir(os.path.join(parent_dir, sub_dir)):
            for fn in glob.glob(os.path.join(parent_dir, sub_dir, file_ext)):
                try:
                    mfccs,mel= extract_feature(fn)
                except Exception as e:
                    print("[Error] extract feature error in %s. %s" % (fn,e))
                    continue
                ext_features = np.hstack([mfccs,mel])
               
                features = np.vstack([features,ext_features])
                # labels = np.append(labels, fn.split('/')[1])
                labels = np.append(labels, label)
            print("extract %s features done" % (sub_dir))
    return np.array(features), np.array(labels, dtype = np.int)







ALLOWED_EXTENSIONS = set(["ogg", "jpg", "jpge"])

def allowed_file(filename):

    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

app = Flask(__name__)


@app.route("/")
def index():
    return "Bienvenido al api"


def verificarExistencia(modelo,cate):
    contenido = os.listdir('./data/'+str(modelo)+"/")    
    for x in contenido:
        if(x.lower()==cate.lower()):
            return True
    return False
                


def writeTrain(exactitud,modelo,F1,recall,Precision):
    book = load_workbook('log.xlsx')
    writer = pd.ExcelWriter('log.xlsx',sheet='Errores', engine='openpyxl') 
    writer.book = book
    pandas.io.formats.excel.header_style = None
    data=len(pd.read_excel("log.xlsx",'Entrenamiento'))+1
    exactitud=pd.DataFrame([str(exactitud)])
    f1_d=pd.DataFrame([str(F1)])
    fecha  = pd.DataFrame([date.today().strftime('%d-%m-%Y')]) 
    hora  = pd.DataFrame([datetime.now().strftime('%H:%M:%S')])
    ip=request.remote_addr
    Precision_v=pd.DataFrame([str(Precision)])
    ipCliente=pd.DataFrame([ip])
    recall_v=pd.DataFrame([recall])
    Modelo=pd.DataFrame([modelo])
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    fecha.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=0)
    hora.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=1)
    ipCliente.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=2) 
    exactitud.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=3)
    f1_d.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=4)
    recall_v.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=5)
    Precision_v.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=6)
    Modelo.to_excel(writer, "Entrenamiento",  index = False,startrow=data,header=None,startcol=7)

    writer.save()

def train(modelo):

    X = np.load('./npys/feat'+str(modelo)+'.npy')
    y = np.load('./npys/label'+str(modelo)+'.npy').ravel()

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.1, random_state=233)
    # Count the number of sub-directories in 'data' class_count =
    class_count=len(next(os.walk('data/'+str(modelo).lower()))[1])

    # Build the Neural Network
    model = Sequential()
    model.add(Conv1D(160, 3, activation='relu', input_shape=(168, 1)))
    model.add(Conv1D(160, 3, activation='relu'))
    model.add(MaxPooling1D(3))
    model.add(Conv1D(128, 3, activation='relu'))
    model.add(Conv1D(128, 3, activation='relu'))
    model.add(GlobalAveragePooling1D())
    model.add(Dropout(0.5))
    model.add(Dense(class_count, activation='softmax'))
    model.compile(loss='categorical_crossentropy', optimizer="rmsprop", metrics=['accuracy',f1_m,precision_m, recall_m])

    # Convert label to onehot
    y_train = keras.utils.to_categorical(y_train, num_classes=class_count)
    y_test = keras.utils.to_categorical(y_test, num_classes=class_count)



    X_train = np.expand_dims(X_train, axis=2)
    X_test = np.expand_dims(X_test, axis=2)

    start = time.time()
    model.fit(X_train, y_train, batch_size=64, epochs=500)#64/500
    loss, acc, f1, precision, recall = model.evaluate(X_test,y_test,verbose=0)

    #print('Recordar: ', format(recall_score(y_test, y_prediccion[1])))

    print("F1: "+str(f1))
    print("Recall: "+str(recall))
    print("Precision: "+str(precision))
    print('Exactitud:', acc)
    print('Tiempo:', (time.time()-start))   
    writeTrain(acc,modelo,f1,recall,precision)
    model.save("./models/"+modelo+".h5")
    CargaModelos()


@app.route("/crearModelo", methods=["POST"])
def crearModelo():
    if request.method == "POST":
        modelo="Model"+str(request.form["modelo"]).lower()
        if verificarModelo(modelo):
            return '{"response":"El modelo ya existe","resp":"False"}'   
        else:
            os.mkdir('./data/'+modelo.lower()+"/")
            return '{"response":"El modelo fue creado correctamente","resp":"True"}'   



@app.route("/entrenar", methods=["POST"])
def entrenar():
    if request.method == "POST":
        modelo="Model"+str(request.form["modelo"])
        features, labels = parse_audio_files('./data/'+modelo+"/")
        np.save('./npys/feat'+modelo+'.npy', features)
        np.save('./npys/label'+modelo+'.npy', labels)
        train(modelo)
        return '{"response":"Modelo entrenado correctamente","resp":"True"}'   


@app.route("/createCategory", methods=["POST"])
def createCategory():
    try:
        if request.method == "POST":
            categoria = request.form["categoria"]
            modelo = request.form["modelo"]
            modelo="Model"+modelo
            resp=verificarModelo(modelo)
            print(resp)
            if resp:
                resp=verificarExistencia(modelo,categoria)
                if resp:
                    return '{"response":"categoria ya existente","resp":"False"}'
                else:
                    try:
                        os.mkdir('./data/'+modelo+"/"+str(categoria.lower()))
                        return '{"response":"La categoria fue creada correctamente pero no estara disponible hasta que se ejecute un entrenamiento","resp":"True"}'
                    except error as e:
                        return '{"response":"Se produjo un error","resp":"False"}'
                    
            else:
                return '{"response":"El modelo seleccionado no existe","resp":"False"}'
    except:
        return '{"response":"Verifique Parametros","resp":"False"}'


            



@app.route("/uploadCategoryFile", methods=["POST"])
def upload_cateogory_file():
    try:
        if request.method == "POST":

            f = request.files["file"]
            
            extension = os.path.splitext(f.filename)[1]
            if extension==".zip":
                categoria = request.form["categoria"]
                modelo = request.form["modelo"]
                modelo="Model"+modelo
                resp=verificarModelo(modelo)
                if resp:    
                    resp=verificarExistencia(modelo,categoria)
                    if resp:
                        archivo_zip = zipfile.ZipFile(f, "r")
                        archivo_zip.extractall(pwd=None, path="./data/"+str(modelo)+"/"+str(categoria.lower()))
                        archivo_zip.close()
                        return '{"response":"Los audios se han subido correctamente, pero no se añadiran al modelo hasta que se ejecute un entrenamiento","resp":"True"}'
                    else:
                        return '{"response":"La categoria no existe","resp":"False"}'
                else:
                    return '{"response":"El modelo seleccionado no existe","resp":"False"}'

            else:
                return '{"response":"Solo se permiten archivos zip","resp":"False"}'
    except:
        return '{"response":"Verifique Parametros","resp":"False"}'








@app.route("/category", methods=["POST"])
def upload_file():
    try:
        if request.method == "POST":
            f = request.files["file"]
            modelo = request.form["modelo"]
            try:

                lista = os.listdir('./data/Model'+str(modelo))
            except:
                writeError("El modelo seleccionado no existe")
                return '{"response":"Verifique los parametros","resp":"False"}'
            extension = os.path.splitext(f.filename)[1]
            if os.path.isfile("./models/Model"+str(modelo)+".h5"):
                if extension==".wav":
                    nombreArchivo=str(randint(0,999999999999999))+".wav"
                    existe=True
                    while existe:
                        if os.path.isfile('predict/'+str(nombreArchivo)):
                            nombreArchivo=str(randint(0,999999999999999))+".wav"
                        else:
                            existe=False
                    nombreArchivo=str(nombreArchivo)+".wav"
                    f.save(os.path.join("predict", nombreArchivo))
                    X_predict, filenames = parse_predict_files(nombreArchivo,'predict')
                    if op.exists("./models/Model"+str(modelo)+".h5"):
                        model = BuscarModelo(str(modelo))
                        X_predict = np.expand_dims(X_predict, axis=2)
                        pred = model.predict(X_predict)
                        for pair in list(zip(filenames, pred)): 
                            nombre,datos=pair
                            maxi=max(datos)
                            array=datos.tolist()
                            pos=array.index(maxi)
                            print(nombre+": "+str(pos+1)+" Exactitud: "+str(maxi))
                            

                            data=len(pd.read_excel("log.xlsx"))+1
                            book = load_workbook('log.xlsx')
                            writer = pd.ExcelWriter('log.xlsx', engine='openpyxl') 
                            writer.book = book
                            pandas.io.formats.excel.header_style = None
                            ip=request.remote_addr
                            #datos log
                            fecha  = pd.DataFrame([date.today().strftime('%d-%m-%Y')])
                            hora  = pd.DataFrame([datetime.now().strftime('%H:%M:%S')])
                            claseid=pd.DataFrame([str(pos+1)])
                            clase=pd.DataFrame([str(lista[pos])])
                            model=pd.DataFrame([str(modelo)])
                            ipCliente=pd.DataFrame([ip])
                            Exactitud=pd.DataFrame([maxi])
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                            fecha.to_excel(writer, "Hoja1",  index = False,startrow=data,header=None,startcol=0) 
                            hora.to_excel(writer, "Hoja1",  index = False,startrow=data,header=None,startcol=1) 
                            claseid.to_excel(writer, "Hoja1",  index = False,startrow=data,header=None,startcol=2) 
                            clase.to_excel(writer, "Hoja1",  index = False,startrow=data,header=None,startcol=3) 
                            Exactitud.to_excel(writer, "Hoja1",  index = False,startrow=data,header=None,startcol=4) 
                            ipCliente.to_excel(writer, "Hoja1",  index = False,startrow=data,header=None,startcol=5) 
                            model.to_excel(writer, "Hoja1",  index = False,startrow=data,header=None,startcol=6) 

                            writer.save()
                            #fin datos log
                        try:
                            remove("predict/"+nombreArchivo)
                        except Exception as e:
                            print("borrado")

                        return '{"categoria":"'+str(pos+1)+'","exactitud":"'+str(maxi)+'","descripcion":"'+str(lista[pos])+'"}'
                    else:
                        writeError("Modelo no existente")
                        return '{"response":"El modelo seleccionado no existe","resp":"False"}'

                else:
                    writeError("Extencion no permitida")
                    return '{"response":"Extencion de archivo no permitida","resp":"False"}'

        else:
            writeError("El modelo seleccionado no existe")
            return '{"response":"El modelo seleccionado no existe","resp":"False"}'
    except ValueError as x:
        print(x)
        return '{"response":"Verifique los parametros","resp":"False"}'



def writeError(error):
    book = load_workbook('log.xlsx')
    writer = pd.ExcelWriter('log.xlsx',sheet='Errores', engine='openpyxl') 
    writer.book = book
    pandas.io.formats.excel.header_style = None
    data=len(pd.read_excel("log.xlsx",'Errores'))+1
    error=pd.DataFrame([str(error)])
    fecha  = pd.DataFrame([date.today().strftime('%d-%m-%Y')]) 
    hora  = pd.DataFrame([datetime.now().strftime('%H:%M:%S')])
    ip=request.remote_addr
    ipCliente=pd.DataFrame([ip])
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    fecha.to_excel(writer, "Errores",  index = False,startrow=data,header=None,startcol=0)
    hora.to_excel(writer, "Errores",  index = False,startrow=data,header=None,startcol=1)
    ipCliente.to_excel(writer, "Errores",  index = False,startrow=data,header=None,startcol=2) 
    error.to_excel(writer, "Errores",  index = False,startrow=data,header=None,startcol=3)
    writer.save()
            
@app.route("/verModelosEntranados", methods=["POST"])
def verModelosEntrenados():
    try:
        if request.method == "POST":
            contenido = os.listdir('./models/')    
            lista=[]
            for e in contenido:
                name=e.replace("Model", "")
                name=name.replace(".H5", "")
                name=name.replace(".h5", "")
                lista.append(name)
            return json.dumps(lista)
    except:
        return '{"response":"Error al listar","resp":"False"}'



if __name__ == "__main__":
    #app.run(ssl_context=('https/cert.pem', 'https/key.pem'))
    app.run()
